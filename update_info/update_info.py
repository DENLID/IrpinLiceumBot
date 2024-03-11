from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import string, datetime
from pytz import timezone
from config import path_ms as path
import json


alphabet_ukr = ["А", "Б", "В", "Г", "Д", "Е", "Є", "Ж", "З", "И"]


def update_info_ms(class_letter, class_number, class_student, present_students, ms_number_hv, ms_students):
    ukraine_time = timezone('Europe/Kiev')
    dt = datetime.datetime.now(ukraine_time)
    dmy = f"{dt.day}.{dt.month}.{dt.year}"

    wb = load_workbook(path)
    
    try:
        xl = wb[f"{dmy}"]
    except:
        wb.create_sheet(f"{dmy}")
        xl = wb[f"{dmy}"]
    
        xl[f"A1"] = "Клас"
        xl[f"B1"] = "Кількість учнів в класі"
        xl[f"C1"] = "Кількість присутніх в класі"
        xl[f"D1"] = "Кількість хворих відсутніх"
        xl[f"E1"] = "Відсутні"

        for l in ["A","B","C"]:
            xl[f"{l}1"].font = xl[f"{l}1"].font.copy(bold=True)
            for n in range(100):
                xl[f'{l}{n+1}'].alignment = Alignment(horizontal='center')
        xl[f"D1"].font = xl[f"D1"].font.copy(bold=True)

    number = ((class_number-1)*11)+alphabet_ukr.index(class_letter)+2

    xl[f"A{number}"] = f"{class_number} - {class_letter}"
    xl[f"B{number}"] = class_student
    xl[f"C{number}"] = present_students
    xl[f"D{number}"] = ms_number_hv
    xl[f"E{number}"] = ms_students
        
    for col in xl.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length+2
        xl.column_dimensions[column].width = adjusted_width
    
    wb.save(path)
    wb.close()